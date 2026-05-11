#!/usr/bin/env python3
"""
Coleta métricas estáticas e dinâmicas do projeto Python/FastAPI.

Gera:
  metrics/report_YYYY-MM-DD_HHMMSS.json  — dados brutos + sumário
  metrics/report_YYYY-MM-DD_HHMMSS.md   — relatório legível
  metrics/report_YYYY-MM-DD_HHMMSS.xlsx — planilha para análise
  metrics/report_YYYY-MM-DD_HHMMSS.png  — gráficos (4 charts)

Uso:
  python scripts/metrics.py
  python scripts/metrics.py --output-dir /caminho/customizado
  python scripts/metrics.py --src app/       # projetos Flask/outros

Requer:
  pip install -r requirements-metrics.txt
"""
import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path


ROOT = Path(__file__).parent.parent
SRC = str(ROOT / "src")


def run(cmd: list[str]) -> tuple[int, str, str]:
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=ROOT)
    return result.returncode, result.stdout, result.stderr


def py(module: str, *args: str) -> list[str]:
    """Roda módulo Python via sys.executable — funciona em qualquer venv."""
    return [sys.executable, "-m", module, *args]


def check_tool(name: str) -> bool:
    code, _, _ = run(py(name, "--version"))
    return code == 0


# ── Coletores ────────────────────────────────────────────────────────────────

def collect_radon_cc(src: str) -> dict:
    """Complexidade ciclomática (McCabe) por função."""
    _, out, err = run(py("radon", "cc", src, "-s", "-a", "-j"))
    try:
        return json.loads(out)
    except json.JSONDecodeError:
        return {"error": err.strip()}


def collect_radon_mi(src: str) -> dict:
    """Índice de manutenibilidade (0–100) por módulo."""
    _, out, err = run(py("radon", "mi", src, "-s", "-j"))
    try:
        return json.loads(out)
    except json.JSONDecodeError:
        return {"error": err.strip()}


def collect_radon_hal(src: str) -> dict:
    """Métricas de Halstead (volume, dificuldade, esforço, bugs estimados)."""
    _, out, err = run(py("radon", "hal", src, "-j"))
    try:
        return json.loads(out)
    except json.JSONDecodeError:
        return {"error": err.strip()}


def collect_pylint(src: str) -> dict:
    """Score de qualidade Pylint (0–10) e issues por categoria."""
    DISABLE = "--disable=C0114,C0115,C0116"

    # Primeira passagem: mensagens estruturadas (JSON)
    _, out_json, _ = run(py("pylint", src, "--output-format=json", DISABLE))
    messages = []
    try:
        messages = json.loads(out_json)
    except json.JSONDecodeError:
        pass

    # Segunda passagem: score (não sai no modo json — requer texto)
    _, out_text, _ = run(py("pylint", src, "--score=yes", "--output-format=text", DISABLE))
    score_line = None
    for line in out_text.splitlines():
        if "Your code has been rated" in line:
            score_line = line.strip()
            break

    return {"messages": messages, "score_line": score_line}


def collect_xenon(src: str) -> dict:
    """Verifica thresholds de complexidade (A=simples … F=caótico)."""
    code, out, err = run(py(
        "xenon",
        "--max-absolute", "C",   # nenhuma função pode passar de C
        "--max-modules", "B",    # nenhum módulo passa de B
        "--max-average", "A",    # média deve ser A
        src,
    ))
    return {
        "passed": code == 0,
        "output": (out or err).strip(),
        "thresholds": {
            "max_absolute": "C",
            "max_modules": "B",
            "max_average": "A",
        },
    }


def collect_coverage(src_module: str) -> dict:
    """Cobertura de testes via pytest-cov."""
    run(py("pytest", f"--cov={src_module}", "--cov-report=json", "-q", "--no-header"))
    coverage_file = ROOT / "coverage.json"
    if not coverage_file.exists():
        return {"error": "coverage.json não encontrado — rode pytest primeiro"}

    data = json.loads(coverage_file.read_text())
    totals = data.get("totals", {})

    # Coleta por arquivo para gráficos
    files = {}
    for fname, fdata in data.get("files", {}).items():
        short = Path(fname).name
        files[short] = round(fdata.get("summary", {}).get("percent_covered", 0), 1)

    return {
        "percent": round(totals.get("percent_covered", 0), 2),
        "covered_lines": totals.get("covered_lines", 0),
        "missing_lines": totals.get("missing_lines", 0),
        "num_statements": totals.get("num_statements", 0),
        "by_file": files,
    }


# ── Sumários ─────────────────────────────────────────────────────────────────

def summarize_cc(cc_data: dict) -> dict:
    complexities = []
    per_file: dict[str, float] = {}

    for filepath, file_items in cc_data.items():
        if not isinstance(file_items, list):
            continue
        file_scores = []
        for item in file_items:
            if isinstance(item, dict):
                complexities.append(item.get("complexity", 0))
                file_scores.append(item.get("complexity", 0))
                for method in item.get("methods", []):
                    complexities.append(method.get("complexity", 0))
                    file_scores.append(method.get("complexity", 0))
        if file_scores:
            per_file[Path(filepath).name] = round(sum(file_scores) / len(file_scores), 2)

    if not complexities:
        return {}

    return {
        "average": round(sum(complexities) / len(complexities), 2),
        "max": max(complexities),
        "min": min(complexities),
        "total_functions": len(complexities),
        "grade": _cc_grade(sum(complexities) / len(complexities)),
        "per_file": per_file,
    }


def _cc_grade(avg: float) -> str:
    if avg <= 5:   return "A (simples)"
    if avg <= 10:  return "B (bem estruturado)"
    if avg <= 15:  return "C (levemente complexo)"
    if avg <= 20:  return "D (complexo)"
    if avg <= 25:  return "E (muito complexo)"
    return "F (instável)"


def summarize_mi(mi_data: dict) -> dict:
    scores = []
    per_file: dict[str, float] = {}

    for fpath, v in mi_data.items():
        if isinstance(v, dict) and "mi" in v:
            scores.append(v["mi"])
            per_file[Path(fpath).name] = round(v["mi"], 2)

    if not scores:
        return {}

    avg = sum(scores) / len(scores)
    return {
        "average": round(avg, 2),
        "min": round(min(scores), 2),
        "max": round(max(scores), 2),
        "grade": _mi_grade(avg),
        "per_file": per_file,
    }


def _mi_grade(mi: float) -> str:
    if mi >= 20: return "Alta manutenibilidade"
    if mi >= 10: return "Manutenibilidade moderada"
    return "Baixa manutenibilidade"


def summarize_hal(hal_data: dict) -> dict:
    """Agrega métricas Halstead — foca em bugs estimados e esforço."""
    total_bugs = 0.0
    total_effort = 0.0
    file_count = 0

    for file_data in hal_data.values():
        if not isinstance(file_data, dict):
            continue
        total = file_data.get("total", {})
        if total:
            total_bugs += total.get("bugs", 0)
            total_effort += total.get("effort", 0)
            file_count += 1

    return {
        "estimated_bugs": round(total_bugs, 4),
        "total_effort": round(total_effort, 2),
        "files_analyzed": file_count,
    }


def summarize_pylint(pylint_data: dict) -> dict:
    messages = pylint_data.get("messages", [])
    by_type: dict[str, int] = {}
    for msg in messages:
        t = msg.get("type", "unknown")
        by_type[t] = by_type.get(t, 0) + 1

    # Extrai score numérico da linha "Your code has been rated at X.XX/10"
    score_num = None
    score_line = pylint_data.get("score_line", "")
    if score_line and "rated at" in score_line:
        try:
            score_num = float(score_line.split("rated at")[1].split("/")[0].strip())
        except (ValueError, IndexError):
            pass

    return {
        "total_issues": len(messages),
        "by_type": by_type,
        "score_line": score_line,
        "score": score_num,
    }


# ── Export: XLSX ──────────────────────────────────────────────────────────────

def generate_xlsx(report: dict, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Cores
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
    HDR_FONT = Font(color="FFFFFF", bold=True)

    def write_table(ws, headers: list, rows: list, start_row: int = 1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(rows, start_row + 1):
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=col, value=val)
                if r % 2 == 0:
                    cell.fill = ALT_FILL
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 28

    # ── Aba 1: Sumário ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sumário"
    cc = report["cyclomatic_complexity"]["summary"]
    mi = report["maintainability_index"]["summary"]
    hal = report["halstead"]["summary"]
    cov = report["test_coverage"]
    pyl = report["pylint"]["summary"]
    xen = report["xenon"]

    rows = [
        ("Gerado em", report["generated_at"]),
        ("Projeto", report["project"]),
        ("", ""),
        ("── Complexidade Ciclomática ──", ""),
        ("CC Média", cc.get("average")),
        ("CC Máxima", cc.get("max")),
        ("CC Mínima", cc.get("min")),
        ("Funções analisadas", cc.get("total_functions")),
        ("Classificação CC", cc.get("grade")),
        ("", ""),
        ("── Manutenibilidade ──", ""),
        ("MI Média", mi.get("average")),
        ("MI Mínima", mi.get("min")),
        ("MI Máxima", mi.get("max")),
        ("Classificação MI", mi.get("grade")),
        ("", ""),
        ("── Halstead ──", ""),
        ("Bugs estimados", hal.get("estimated_bugs")),
        ("Esforço total", hal.get("total_effort")),
        ("Arquivos analisados", hal.get("files_analyzed")),
        ("", ""),
        ("── Cobertura de Testes ──", ""),
        ("Cobertura %", cov.get("percent")),
        ("Linhas cobertas", cov.get("covered_lines")),
        ("Linhas faltando", cov.get("missing_lines")),
        ("Total statements", cov.get("num_statements")),
        ("", ""),
        ("── Pylint ──", ""),
        ("Score", pyl.get("score")),
        ("Total issues", pyl.get("total_issues")),
        ("", ""),
        ("── Xenon ──", ""),
        ("Resultado", "✅ Passou" if xen.get("passed") else "❌ Falhou"),
        ("Max absoluto", xen["thresholds"]["max_absolute"]),
        ("Max módulo", xen["thresholds"]["max_modules"]),
        ("Média máxima", xen["thresholds"]["max_average"]),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    for r, (label, value) in enumerate(rows, 1):
        la = ws.cell(row=r, column=1, value=label)
        vb = ws.cell(row=r, column=2, value=value)
        if label.startswith("──"):
            la.font = Font(bold=True)
        if label == "Resultado":
            vb.fill = PASS_FILL if xen.get("passed") else FAIL_FILL

    # ── Aba 2: CC por arquivo ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("CC por Arquivo")
    per_file = cc.get("per_file", {})
    write_table(ws2,
        ["Arquivo", "CC Média"],
        sorted(per_file.items(), key=lambda x: x[1], reverse=True),
    )

    # ── Aba 3: MI por arquivo ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("MI por Arquivo")
    mi_per_file = mi.get("per_file", {})
    write_table(ws3,
        ["Arquivo", "Índice de Manutenibilidade"],
        sorted(mi_per_file.items(), key=lambda x: x[1]),
    )

    # ── Aba 4: Cobertura por arquivo ──────────────────────────────────────────
    ws4 = wb.create_sheet("Cobertura por Arquivo")
    cov_files = cov.get("by_file", {})
    cov_rows = sorted(cov_files.items(), key=lambda x: x[1])
    write_table(ws4, ["Arquivo", "Cobertura %"], cov_rows)
    # Colore linhas baixas/altas
    for r, (_, pct) in enumerate(cov_rows, 2):
        fill = PASS_FILL if pct >= 80 else FAIL_FILL if pct < 50 else ALT_FILL
        ws4.cell(row=r, column=2).fill = fill

    # ── Aba 5: Pylint issues ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Pylint Issues")
    by_type = pyl.get("by_type", {})
    write_table(ws5,
        ["Tipo", "Quantidade"],
        sorted(by_type.items(), key=lambda x: x[1], reverse=True),
    )

    # ── Aba 6: Dados brutos (JSON) ────────────────────────────────────────────
    ws6 = wb.create_sheet("JSON Bruto")
    ws6.cell(row=1, column=1, value="JSON completo para ingestão em IA:")
    ws6.cell(row=1, column=1).font = Font(bold=True)
    raw_json = json.dumps(report, indent=2, ensure_ascii=False)
    for i, line in enumerate(raw_json.splitlines(), 2):
        ws6.cell(row=i, column=1, value=line)
    ws6.column_dimensions["A"].width = 120

    wb.save(path)


# ── Export: PNG individuais (matplotlib) ─────────────────────────────────────

def _fig_save(fig, path: Path) -> None:
    """Salva figura e fecha."""
    import matplotlib.pyplot as plt
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def generate_charts(report: dict, base: Path) -> dict[str, Path]:
    """
    Gera 4 PNGs individuais. Retorna dict {slug: Path}.
    base = output_dir / f"report_{ts}"  (sem extensão)
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    BG = "#F8F9FA"
    AX_BG = "#FAFAFA"
    cc = report["cyclomatic_complexity"]["summary"]
    mi = report["maintainability_index"]["summary"]
    cov = report["test_coverage"]
    pyl = report["pylint"]["summary"]
    paths: dict[str, Path] = {}

    # ── 1: Complexidade Ciclomática por arquivo ───────────────────────────────
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(AX_BG)
    per_file = cc.get("per_file", {})
    if per_file:
        items = sorted(per_file.items(), key=lambda x: x[1], reverse=True)[:20]
        files, vals = zip(*items)
        colors = ["#C00000" if v > 10 else "#ED7D31" if v > 5 else "#70AD47" for v in vals]
        bars = ax.barh(files, vals, color=colors, edgecolor="white", linewidth=0.5)
        ax.axvline(x=5,  color="#70AD47", linestyle="--", alpha=0.7, label="Limite A (≤5)")
        ax.axvline(x=10, color="#ED7D31", linestyle="--", alpha=0.7, label="Limite B (≤10)")
        ax.legend(fontsize=9)
        for bar, val in zip(bars, vals):
            ax.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
                    f"{val}", va="center", fontsize=8)
    else:
        ax.text(0.5, 0.5, "Sem dados", ha="center", va="center", transform=ax.transAxes)
    ax.set_xlabel("CC Média por Arquivo")
    ax.set_title(
        f"Complexidade Ciclomática — média geral: {cc.get('average', 'N/A')} ({cc.get('grade', '')})",
        fontweight="bold", pad=12,
    )
    p = Path(f"{base}_cc.png")
    _fig_save(fig, p)
    paths["cc"] = p

    # ── 2: Manutenibilidade por arquivo ──────────────────────────────────────
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(AX_BG)
    mi_per_file = mi.get("per_file", {})
    if mi_per_file:
        items = sorted(mi_per_file.items(), key=lambda x: x[1])[:20]
        files, vals = zip(*items)
        colors = ["#C00000" if v < 10 else "#ED7D31" if v < 20 else "#70AD47" for v in vals]
        bars = ax.barh(files, vals, color=colors, edgecolor="white", linewidth=0.5)
        ax.axvline(x=20, color="#70AD47", linestyle="--", alpha=0.7, label="Alta manutenib. (≥20)")
        ax.axvline(x=10, color="#ED7D31", linestyle="--", alpha=0.7, label="Moderada (≥10)")
        ax.set_xlim(0, 105)
        ax.legend(fontsize=9)
        for bar, val in zip(bars, vals):
            ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}", va="center", fontsize=8)
    else:
        ax.text(0.5, 0.5, "Sem dados", ha="center", va="center", transform=ax.transAxes)
    ax.set_xlabel("Índice de Manutenibilidade (0–100)")
    ax.set_title(
        f"Manutenibilidade — média: {mi.get('average', 'N/A')} ({mi.get('grade', '')})",
        fontweight="bold", pad=12,
    )
    p = Path(f"{base}_mi.png")
    _fig_save(fig, p)
    paths["mi"] = p

    # ── 3: Cobertura por arquivo ──────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(AX_BG)
    cov_pct = cov.get("percent", 0)
    cov_files = cov.get("by_file", {})
    if cov_files:
        items = sorted(cov_files.items(), key=lambda x: x[1])[:20]
        files, vals = zip(*items)
        colors = ["#C00000" if v < 50 else "#ED7D31" if v < 80 else "#70AD47" for v in vals]
        bars = ax.barh(files, vals, color=colors, edgecolor="white", linewidth=0.5)
        ax.axvline(x=80, color="#70AD47", linestyle="--", alpha=0.7, label="Meta 80%")
        ax.set_xlim(0, 108)
        ax.legend(fontsize=9)
        for bar, val in zip(bars, vals):
            ax.text(min(bar.get_width() + 1, 105), bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}%", va="center", fontsize=8)
    else:
        # Semicírculo gauge quando sem dados por arquivo
        theta = np.linspace(0, np.pi, 200)
        ax.plot(np.cos(theta), np.sin(theta), color="#E0E0E0", linewidth=28, solid_capstyle="round")
        end_theta = np.linspace(0, cov_pct / 100 * np.pi, 200)
        color = "#70AD47" if cov_pct >= 80 else "#ED7D31" if cov_pct >= 50 else "#C00000"
        ax.plot(np.cos(end_theta), np.sin(end_theta), color=color, linewidth=28, solid_capstyle="round")
        ax.text(0, 0.1, f"{cov_pct}%", ha="center", va="center", fontsize=40,
                fontweight="bold", color=color)
        ax.set_xlim(-1.3, 1.3)
        ax.set_ylim(-0.2, 1.3)
        ax.axis("off")
    ax.set_xlabel("Cobertura (%)" if cov_files else "")
    ax.set_title(f"Cobertura de Testes — total: {cov_pct}%", fontweight="bold", pad=12)
    p = Path(f"{base}_coverage.png")
    _fig_save(fig, p)
    paths["coverage"] = p

    # ── 4: Pylint — pizza + score ─────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(8, 7))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG)
    by_type = pyl.get("by_type", {})
    score = pyl.get("score")
    if by_type:
        TYPE_COLORS = {
            "convention": "#4472C4", "refactor": "#ED7D31",
            "warning": "#FFC000",   "error": "#C00000", "fatal": "#7030A0",
        }
        labels, sizes = zip(*sorted(by_type.items(), key=lambda x: x[1], reverse=True))
        pie_colors = [TYPE_COLORS.get(l, "#A9A9A9") for l in labels]
        _, texts, autotexts = ax.pie(
            sizes, labels=labels, colors=pie_colors,
            autopct=lambda p: f"{p:.0f}%\n({int(round(p * sum(sizes) / 100))})",
            startangle=90, wedgeprops={"edgecolor": "white", "linewidth": 2},
            pctdistance=0.75,
        )
        for at in autotexts:
            at.set_fontsize(9)
        score_str = f"Score: {score:.2f}/10" if score is not None else ""
        ax.set_title(
            f"Pylint — {pyl.get('total_issues', 0)} issues\n{score_str}",
            fontweight="bold", pad=14,
        )
    else:
        score_color = "#70AD47" if (score or 0) >= 8 else "#ED7D31" if (score or 0) >= 6 else "#C00000"
        ax.text(0.5, 0.5, f"{score or 'N/A'}/10", ha="center", va="center",
                transform=ax.transAxes, fontsize=52, fontweight="bold", color=score_color)
        ax.set_title("Pylint Score", fontweight="bold", pad=14)
        ax.axis("off")
    p = Path(f"{base}_pylint.png")
    _fig_save(fig, p)
    paths["pylint"] = p

    return paths


# ── Export: HTML standalone ───────────────────────────────────────────────────

def generate_html(report: dict, chart_paths: dict[str, Path], path: Path) -> None:
    """HTML standalone com gráficos embutidos em base64. Abre direto no browser."""
    import base64

    def img_b64(p: Path) -> str:
        return base64.b64encode(p.read_bytes()).decode()

    cc = report["cyclomatic_complexity"]["summary"]
    mi = report["maintainability_index"]["summary"]
    hal = report["halstead"]["summary"]
    cov = report["test_coverage"]
    pyl = report["pylint"]["summary"]
    xen = report["xenon"]
    ts = report["generated_at"]

    def kv_table(rows: list[tuple]) -> str:
        trs = "".join(
            f'<tr><td class="label">{k}</td><td class="value">{v if v is not None else "N/A"}</td></tr>'
            for k, v in rows
        )
        return f'<table class="kv">{trs}</table>'

    def badge(ok: bool) -> str:
        cls = "pass" if ok else "fail"
        txt = "✅ Passou" if ok else "❌ Falhou"
        return f'<span class="badge {cls}">{txt}</span>'

    charts_html = "".join(
        f'''<div class="chart-block">
              <img src="data:image/png;base64,{img_b64(p)}" alt="{slug}">
           </div>'''
        for slug, p in chart_paths.items()
        if p.exists()
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Métricas — {ts}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
          background: #F0F2F5; color: #1a1a2e; }}
  header {{ background: #1F3864; color: #fff; padding: 24px 32px; }}
  header h1 {{ font-size: 1.4rem; font-weight: 700; }}
  header p  {{ font-size: 0.85rem; opacity: 0.75; margin-top: 4px; }}
  main {{ max-width: 1200px; margin: 32px auto; padding: 0 24px; }}
  h2 {{ font-size: 1rem; font-weight: 700; color: #1F3864;
        border-left: 4px solid #1F3864; padding-left: 10px;
        margin: 32px 0 12px; }}
  .cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
             gap: 16px; margin-bottom: 32px; }}
  .card {{ background: #fff; border-radius: 10px; padding: 20px;
            box-shadow: 0 1px 4px rgba(0,0,0,.08); text-align: center; }}
  .card .num {{ font-size: 2rem; font-weight: 700; color: #1F3864; }}
  .card .lbl {{ font-size: 0.75rem; color: #666; margin-top: 4px; }}
  .card .sub {{ font-size: 0.8rem; color: #888; margin-top: 2px; }}
  table.kv {{ width: 100%; border-collapse: collapse; background: #fff;
               border-radius: 8px; overflow: hidden;
               box-shadow: 0 1px 4px rgba(0,0,0,.06); margin-bottom: 24px; }}
  table.kv td {{ padding: 8px 14px; font-size: 0.875rem; border-bottom: 1px solid #F0F2F5; }}
  table.kv td.label {{ color: #555; width: 50%; }}
  table.kv td.value {{ font-weight: 600; }}
  .badge {{ display: inline-block; padding: 3px 10px; border-radius: 20px;
             font-size: 0.8rem; font-weight: 600; }}
  .badge.pass {{ background: #C6EFCE; color: #276221; }}
  .badge.fail {{ background: #FFC7CE; color: #9C0006; }}
  .charts {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(460px, 1fr));
              gap: 24px; margin-bottom: 32px; }}
  .chart-block {{ background: #fff; border-radius: 10px; overflow: hidden;
                   box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  .chart-block img {{ width: 100%; height: auto; display: block; }}
  footer {{ text-align: center; padding: 32px; color: #aaa; font-size: 0.8rem; }}
</style>
</head>
<body>
<header>
  <h1>📊 Relatório de Métricas</h1>
  <p>Projeto: {report["project"]} &nbsp;·&nbsp; Gerado em: {ts}</p>
</header>
<main>

<h2>Resumo</h2>
<div class="cards">
  <div class="card">
    <div class="num">{cc.get("average", "N/A")}</div>
    <div class="lbl">CC Média</div>
    <div class="sub">{cc.get("grade", "")}</div>
  </div>
  <div class="card">
    <div class="num">{mi.get("average", "N/A")}</div>
    <div class="lbl">MI Média</div>
    <div class="sub">{mi.get("grade", "")}</div>
  </div>
  <div class="card">
    <div class="num">{cov.get("percent", "N/A")}%</div>
    <div class="lbl">Cobertura</div>
    <div class="sub">{cov.get("covered_lines", 0)} / {cov.get("num_statements", 0)} linhas</div>
  </div>
  <div class="card">
    <div class="num">{pyl.get("score") or "N/A"}</div>
    <div class="lbl">Pylint Score</div>
    <div class="sub">{pyl.get("total_issues", 0)} issues</div>
  </div>
  <div class="card">
    <div class="num">{hal.get("estimated_bugs", "N/A")}</div>
    <div class="lbl">Bugs estimados</div>
    <div class="sub">Halstead</div>
  </div>
  <div class="card">
    <div class="num">{"✅" if xen.get("passed") else "❌"}</div>
    <div class="lbl">Xenon</div>
    <div class="sub">max-absolute C | avg A</div>
  </div>
</div>

<h2>Gráficos</h2>
<div class="charts">{charts_html}</div>

<h2>Complexidade Ciclomática</h2>
{kv_table([
    ("Média", cc.get("average")), ("Máxima", cc.get("max")), ("Mínima", cc.get("min")),
    ("Funções analisadas", cc.get("total_functions")), ("Classificação", cc.get("grade")),
])}

<h2>Manutenibilidade (MI)</h2>
{kv_table([
    ("Média", mi.get("average")), ("Mínima", mi.get("min")),
    ("Máxima", mi.get("max")), ("Classificação", mi.get("grade")),
])}

<h2>Halstead</h2>
{kv_table([
    ("Bugs estimados", hal.get("estimated_bugs")),
    ("Esforço total", hal.get("total_effort")),
    ("Arquivos analisados", hal.get("files_analyzed")),
])}

<h2>Cobertura de Testes</h2>
{kv_table([
    ("Total", f"{cov.get('percent', 'N/A')}%"),
    ("Linhas cobertas", cov.get("covered_lines")),
    ("Linhas faltando", cov.get("missing_lines")),
    ("Total statements", cov.get("num_statements")),
])}

<h2>Pylint</h2>
{kv_table(
    [("Score", pyl.get("score_line") or "N/A"), ("Total issues", pyl.get("total_issues"))]
    + [(f"Issues — {t}", n) for t, n in pyl.get("by_type", {}).items()]
)}

<h2>Xenon — Thresholds</h2>
{kv_table([
    ("Resultado", badge(xen.get("passed", False))),
    ("Max absoluto", xen["thresholds"]["max_absolute"]),
    ("Max módulo", xen["thresholds"]["max_modules"]),
    ("Média máxima", xen["thresholds"]["max_average"]),
])}

</main>
<footer>Gerado por scripts/metrics.py · KaiserInc</footer>
</body>
</html>"""

    path.write_text(html, encoding="utf-8")


# ── Relatório Markdown ────────────────────────────────────────────────────────

def generate_markdown(report: dict) -> str:
    ts = report["generated_at"]
    cc = report["cyclomatic_complexity"]["summary"]
    mi = report["maintainability_index"]["summary"]
    hal = report["halstead"]["summary"]
    cov = report["test_coverage"]
    pyl = report["pylint"]["summary"]
    xen = report["xenon"]

    lines = [
        f"# Relatório de Métricas — {ts}",
        "",
        "> Gerado por `scripts/metrics.py`. Utilize para comparativos acadêmicos.",
        "",
        "---",
        "",
        "## 1. Complexidade Ciclomática (McCabe)",
        "",
        "Mede o número de caminhos independentes em cada função.",
        "Referência: A ≤ 5 | B ≤ 10 | C ≤ 15 | D ≤ 20 | E ≤ 25 | F > 25",
        "",
        f"| Métrica | Valor |",
        f"|---------|-------|",
        f"| Média   | **{cc.get('average', 'N/A')}** |",
        f"| Máxima  | {cc.get('max', 'N/A')} |",
        f"| Mínima  | {cc.get('min', 'N/A')} |",
        f"| Funções analisadas | {cc.get('total_functions', 0)} |",
        f"| Classificação | **{cc.get('grade', 'N/A')}** |",
        "",
        "---",
        "",
        "## 2. Índice de Manutenibilidade (MI)",
        "",
        "Escala 0–100. ≥ 20 = alta manutenibilidade.",
        "",
        f"| Métrica | Valor |",
        f"|---------|-------|",
        f"| Média   | **{mi.get('average', 'N/A')}** |",
        f"| Mínimo  | {mi.get('min', 'N/A')} |",
        f"| Máximo  | {mi.get('max', 'N/A')} |",
        f"| Classificação | **{mi.get('grade', 'N/A')}** |",
        "",
        "---",
        "",
        "## 3. Métricas de Halstead",
        "",
        "Derivadas de operadores/operandos. Estima complexidade e bugs.",
        "",
        f"| Métrica | Valor |",
        f"|---------|-------|",
        f"| Bugs estimados | **{hal.get('estimated_bugs', 'N/A')}** |",
        f"| Esforço total  | {hal.get('total_effort', 'N/A')} |",
        f"| Arquivos       | {hal.get('files_analyzed', 0)} |",
        "",
        "---",
        "",
        "## 4. Cobertura de Testes",
        "",
        f"| Métrica | Valor |",
        f"|---------|-------|",
        f"| Cobertura total | **{cov.get('percent', 'N/A')}%** |",
        f"| Linhas cobertas | {cov.get('covered_lines', 'N/A')} |",
        f"| Linhas faltando | {cov.get('missing_lines', 'N/A')} |",
        f"| Total de statements | {cov.get('num_statements', 'N/A')} |",
        "",
        "---",
        "",
        "## 5. Pylint",
        "",
        f"| Métrica | Valor |",
        f"|---------|-------|",
        f"| Total de issues | **{pyl.get('total_issues', 0)}** |",
    ]

    for t, count in pyl.get("by_type", {}).items():
        lines.append(f"| Issues tipo `{t}` | {count} |")

    lines += [
        f"| Score | {pyl.get('score_line') or 'N/A'} |",
        "",
        "---",
        "",
        "## 6. Xenon — Thresholds de Complexidade",
        "",
        f"| Threshold | Valor |",
        f"|-----------|-------|",
        f"| Max absoluto | {xen['thresholds']['max_absolute']} |",
        f"| Max módulo   | {xen['thresholds']['max_modules']} |",
        f"| Média máxima | {xen['thresholds']['max_average']} |",
        f"| Resultado    | {'✅ Passou' if xen.get('passed') else '❌ Falhou'} |",
        "",
        "```",
        xen.get("output", ""),
        "```",
        "",
        "---",
        "",
        "## Arquivos gerados",
        "",
        "| Arquivo | Conteúdo |",
        "|---------|----------|",
        f"| `report_{ts}.json`          | Dados brutos completos |",
        f"| `report_{ts}.md`            | Este relatório |",
        f"| `report_{ts}.xlsx`          | Planilha com 6 abas (para análise com IA) |",
        f"| `report_{ts}.html`          | Relatório visual standalone (abre no browser) |",
        f"| `report_{ts}_cc.png`        | Gráfico: Complexidade Ciclomática |",
        f"| `report_{ts}_mi.png`        | Gráfico: Manutenibilidade |",
        f"| `report_{ts}_coverage.png`  | Gráfico: Cobertura de Testes |",
        f"| `report_{ts}_pylint.png`    | Gráfico: Pylint Issues |",
    ]

    return "\n".join(lines)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Coleta métricas do projeto")
    parser.add_argument("--output-dir", default=str(ROOT / "metrics"), help="Diretório de saída")
    parser.add_argument("--src", default=None, help="Diretório fonte (padrão: src/). Use 'app/' para Flask.")
    args = parser.parse_args()

    src = str(ROOT / args.src) if args.src else SRC
    src_module = args.src.rstrip("/") if args.src else "src"

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Verifica dependências
    missing = [t for t in ["radon", "pylint", "xenon"] if not check_tool(t)]
    if missing:
        print(f"❌ Ferramentas não encontradas: {', '.join(missing)}")
        print("   Execute: make metrics-install")
        sys.exit(1)

    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    print(f"📊 Coletando métricas — {ts}\n")

    print("  → Complexidade ciclomática (radon cc)...")
    cc_raw = collect_radon_cc(src)

    print("  → Índice de manutenibilidade (radon mi)...")
    mi_raw = collect_radon_mi(src)

    print("  → Métricas Halstead (radon hal)...")
    hal_raw = collect_radon_hal(src)

    print("  → Pylint...")
    pylint_raw = collect_pylint(src)

    print("  → Xenon (thresholds)...")
    xenon_data = collect_xenon(src)

    print("  → Cobertura de testes (pytest-cov)...")
    coverage_data = collect_coverage(src_module)

    report = {
        "generated_at": ts,
        "project": "python-fastapi-boilerplate",
        "cyclomatic_complexity": {
            "raw": cc_raw,
            "summary": summarize_cc(cc_raw),
        },
        "maintainability_index": {
            "raw": mi_raw,
            "summary": summarize_mi(mi_raw),
        },
        "halstead": {
            "raw": hal_raw,
            "summary": summarize_hal(hal_raw),
        },
        "pylint": {
            "raw": pylint_raw,
            "summary": summarize_pylint(pylint_raw),
        },
        "xenon": xenon_data,
        "test_coverage": coverage_data,
    }

    base = output_dir / f"report_{ts}"

    print("  → Gerando JSON...")
    json_path = Path(f"{base}.json")
    json_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))

    print("  → Gerando Markdown...")
    md_path = Path(f"{base}.md")
    md_path.write_text(generate_markdown(report))

    print("  → Gerando XLSX...")
    xlsx_path = Path(f"{base}.xlsx")
    generate_xlsx(report, xlsx_path)

    print("  → Gerando gráficos PNG (individuais)...")
    chart_paths = generate_charts(report, base)

    print("  → Gerando HTML standalone...")
    html_path = Path(f"{base}.html")
    generate_html(report, chart_paths, html_path)

    print(f"\n✅ Relatórios gerados:")
    print(f"   JSON  → {json_path}")
    print(f"   MD    → {md_path}")
    print(f"   XLSX  → {xlsx_path}")
    print(f"   HTML  → {html_path}")
    for slug, p in chart_paths.items():
        print(f"   PNG [{slug:8s}] → {p}")

    # Sumário rápido no terminal
    cc_s = report["cyclomatic_complexity"]["summary"]
    mi_s = report["maintainability_index"]["summary"]
    cov_s = report["test_coverage"]

    print(f"\n📋 Sumário:")
    print(f"   CC média:        {cc_s.get('average', 'N/A')} ({cc_s.get('grade', '')})")
    print(f"   MI média:        {mi_s.get('average', 'N/A')} ({mi_s.get('grade', '')})")
    print(f"   Cobertura:       {cov_s.get('percent', 'N/A')}%")
    print(f"   Xenon:           {'✅' if xenon_data.get('passed') else '❌'}")
    print(f"   Pylint issues:   {report['pylint']['summary'].get('total_issues', 0)}")
    pyl_score = report['pylint']['summary'].get('score')
    if pyl_score is not None:
        print(f"   Pylint score:    {pyl_score:.2f}/10")


if __name__ == "__main__":
    main()
